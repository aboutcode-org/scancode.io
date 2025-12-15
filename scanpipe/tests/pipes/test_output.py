# SPDX-License-Identifier: Apache-2.0
#
# http://nexb.com and https://github.com/nexB/scancode.io
# The ScanCode.io software is licensed under the Apache License version 2.0.
# Data generated with ScanCode.io is provided as-is without warranties.
# ScanCode is a trademark of nexB Inc.
#
# You may not use this software except in compliance with the License.
# You may obtain a copy of the License at: http://apache.org/licenses/LICENSE-2.0
# Unless required by applicable law or agreed to in writing, software distributed
# under the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR
# CONDITIONS OF ANY KIND, either express or implied. See the License for the
# specific language governing permissions and limitations under the License.
#
# Data Generated with ScanCode.io is provided on an "AS IS" BASIS, WITHOUT WARRANTIES
# OR CONDITIONS OF ANY KIND, either express or implied. No content created from
# ScanCode.io should be considered or used as legal advice. Consult an Attorney
# for any legal advice.
#
# ScanCode.io is a free software code scanning tool from nexB Inc. and others.
# Visit https://github.com/nexB/scancode.io for support and download.

import collections
import io
import json
import shutil
import tempfile
import uuid
import zipfile
from dataclasses import dataclass
from pathlib import Path
from unittest import mock

from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase

import openpyxl
import xlsxwriter
from licensedcode.cache import get_licensing
from lxml import etree
from scancode_config import __version__ as scancode_toolkit_version

from scanpipe import pipes
from scanpipe.models import CodebaseResource
from scanpipe.models import Project
from scanpipe.pipes import flag
from scanpipe.pipes import output
from scanpipe.tests import FIXTURES_REGEN
from scanpipe.tests import make_dependency
from scanpipe.tests import make_message
from scanpipe.tests import make_package
from scanpipe.tests import make_project
from scanpipe.tests import make_resource_file
from scanpipe.tests import mocked_now
from scanpipe.tests import package_data1


def make_config_directory(project):
    """Make and return the `project` config directory."""
    config_directory = project.codebase_path / settings.SCANCODEIO_CONFIG_DIR
    config_directory.mkdir(exist_ok=True)
    return config_directory


class ScanPipeOutputPipesTest(TestCase):
    data = Path(__file__).parent.parent / "data"

    def assertResultsEqual(self, expected_file, results, regen=FIXTURES_REGEN):
        """Set `regen` to True to regenerate the expected results."""
        if regen:
            expected_file.write_text(results)

        expected_data = expected_file.read_text()
        self.assertEqual(expected_data, results)

    def test_scanpipe_pipes_outputs_queryset_to_csv_file(self):
        project1 = make_project(name="Analysis")
        codebase_resource = CodebaseResource.objects.create(
            project=project1,
            path="filename.ext",
        )
        codebase_resource.create_and_add_package(package_data1)

        queryset = project1.discoveredpackages.all()
        fieldnames = ["purl", "name", "version"]

        output_file_path = project1.get_output_file_path("packages", "csv")
        with output_file_path.open("w") as output_file:
            output.queryset_to_csv_file(queryset, fieldnames, output_file)

        expected = [
            "purl,name,version\n",
            "pkg:deb/debian/adduser@3.118?arch=all,adduser,3.118\n",
        ]
        with output_file_path.open() as f:
            self.assertEqual(expected, f.readlines())

        queryset = project1.codebaseresources.all()
        fieldnames = ["for_packages", "path"]
        output_file_path = project1.get_output_file_path("resources", "csv")
        with output_file_path.open("w") as output_file:
            output.queryset_to_csv_file(queryset, fieldnames, output_file)

        package_uid = "pkg:deb/debian/adduser@3.118?uuid=610bed29-ce39-40e7-92d6-fd8b"
        expected = [
            "for_packages,path\n",
            f"['{package_uid}'],filename.ext\n",
        ]
        with output_file_path.open() as f:
            self.assertEqual(expected, f.readlines())

    def test_scanpipe_pipes_outputs_queryset_to_csv_stream(self):
        project1 = make_project(name="Analysis")
        codebase_resource = CodebaseResource.objects.create(
            project=project1,
            path="filename.ext",
        )
        codebase_resource.create_and_add_package(package_data1)

        queryset = project1.discoveredpackages.all()
        fieldnames = ["purl", "name", "version"]

        output_file = project1.get_output_file_path("packages", "csv")
        with output_file.open("w") as output_stream:
            generator = output.queryset_to_csv_stream(
                queryset, fieldnames, output_stream
            )
            collections.deque(generator, maxlen=0)  # Exhaust the generator

        expected = [
            "purl,name,version\n",
            "pkg:deb/debian/adduser@3.118?arch=all,adduser,3.118\n",
        ]
        with output_file.open() as f:
            self.assertEqual(expected, f.readlines())

        queryset = project1.codebaseresources.all()
        fieldnames = ["for_packages", "path"]
        output_file = project1.get_output_file_path("resources", "csv")
        with output_file.open("w") as output_stream:
            generator = output.queryset_to_csv_stream(
                queryset, fieldnames, output_stream
            )
            collections.deque(generator, maxlen=0)  # Exhaust the generator

        output.queryset_to_csv_stream(queryset, fieldnames, output_file)
        package_uid = "pkg:deb/debian/adduser@3.118?uuid=610bed29-ce39-40e7-92d6-fd8b"
        expected = [
            "for_packages,path\n",
            f"['{package_uid}'],filename.ext\n",
        ]
        with output_file.open() as f:
            self.assertEqual(expected, f.readlines())

    @mock.patch("scanpipe.pipes.datetime", mocked_now)
    def test_scanpipe_pipes_outputs_to_csv(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")

        with self.assertNumQueries(7):
            output_files = output.to_csv(project=project)

        csv_files = [
            "codebaseresource-2010-10-10-10-10-10.csv",
            "discovereddependency-2010-10-10-10-10-10.csv",
            "discoveredpackage-2010-10-10-10-10-10.csv",
            "projectmessage-2010-10-10-10-10-10.csv",
        ]

        for csv_file in csv_files:
            self.assertIn(csv_file, project.output_root)
            self.assertIn(csv_file, [f.name for f in output_files])

    def test_scanpipe_pipes_outputs_to_json(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")
        make_message(project, model="resource", description="Error1")
        make_message(project, model="package", description="Error2")

        output_file = output.to_json(project=project)
        self.assertIn(output_file.name, project.output_root)

        with output_file.open() as f:
            results = json.loads(f.read())

        expected = ["dependencies", "files", "headers", "packages", "relations"]
        self.assertEqual(expected, sorted(results.keys()))

        self.assertEqual(1, len(results["headers"]))
        self.assertEqual(18, len(results["files"]))
        self.assertEqual(2, len(results["packages"]))
        self.assertEqual(4, len(results["dependencies"]))
        self.assertEqual(2, len(results["headers"][0]["messages"]))

        self.assertEqual("scanpipe", results["headers"][0]["tool_name"])
        expected = [f"pkg:pypi/scancode-toolkit@{scancode_toolkit_version}"]
        self.assertEqual(expected, results["headers"][0]["other_tools"])

        self.assertIn("compliance_alert", results["files"][0])

        # Make sure the output can be generated even if the work_directory was wiped
        shutil.rmtree(project.work_directory)
        with self.assertNumQueries(10):
            output_file = output.to_json(project=project)
        self.assertIn(output_file.name, project.output_root)

    def test_scanpipe_pipes_outputs_to_xlsx(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})

        project = Project.objects.get(name="asgiref")
        make_message(project, description="Error")
        make_resource_file(
            project=project, path="path/file1.ext", status=flag.REQUIRES_REVIEW
        )

        with self.assertNumQueries(12):
            output_file = output.to_xlsx(project=project)
        self.assertIn(output_file.name, project.output_root)

        # Make sure the output can be generated even if the work_directory was wiped
        shutil.rmtree(project.work_directory)
        output_file = output.to_xlsx(project=project)
        self.assertIn(output_file.name, project.output_root)

        workbook = openpyxl.load_workbook(output_file, read_only=True, data_only=True)
        expected_sheet_names = [
            "PACKAGES",
            "DEPENDENCIES",
            "RESOURCES",
            "RELATIONS",
            "MESSAGES",
            "TODOS",
        ]
        self.assertEqual(expected_sheet_names, workbook.sheetnames)

    def test_scanpipe_pipes_outputs_get_xlsx_report(self):
        project_qs = None
        model_short_name = None

        expected_message = "None is not valid."
        with self.assertRaisesMessage(ValueError, expected_message):
            output.get_xlsx_report(project_qs, model_short_name)

        model_short_name = "package"
        expected_message = "'NoneType' object is not iterable"
        with self.assertRaisesMessage(TypeError, expected_message):
            output.get_xlsx_report(project_qs, model_short_name)

        make_project()
        make_project()
        project_qs = Project.objects.all()
        output_file = output.get_xlsx_report(project_qs, model_short_name)

        self.assertIsInstance(output_file, io.BytesIO)
        workbook = openpyxl.load_workbook(output_file, read_only=True, data_only=True)
        expected_sheet_names = [
            "PACKAGES",
        ]
        self.assertEqual(expected_sheet_names, workbook.sheetnames)

        model_short_name = "todo"
        output_file = output.get_xlsx_report(project_qs, model_short_name)
        workbook = openpyxl.load_workbook(output_file, read_only=True, data_only=True)
        expected_sheet_names = [
            "TODOS",
        ]
        self.assertEqual(expected_sheet_names, workbook.sheetnames)

    def test_scanpipe_pipes_outputs_get_xlsx_fields_order(self):
        output_file = output.to_xlsx(project=make_project())
        workbook = openpyxl.load_workbook(output_file, read_only=True, data_only=True)

        self.assertIn("RESOURCES", workbook.sheetnames)
        resources_sheet = workbook["RESOURCES"]
        headers = [cell.value for cell in next(resources_sheet.iter_rows())]

        expected_order = [
            "path",
            "type",
            "name",
            "status",
            "for_packages",
            "tag",
            "extension",
            "size",
            "mime_type",
            "file_type",
            "programming_language",
            "detected_license_expression",
            "detected_license_expression_spdx",
            "percentage_of_license_text",
            "copyrights",
            "holders",
            "authors",
            "emails",
            "urls",
            "md5",
            "sha1",
            "sha256",
            "sha512",
            "sha1_git",
            "is_binary",
            "is_text",
            "is_archive",
            "is_media",
            "is_legal",
            "is_manifest",
            "is_readme",
            "is_top_level",
            "is_key_file",
            "xlsx_errors",
        ]
        self.assertEqual(expected_order, headers)

    def test_scanpipe_pipes_outputs_vulnerability_as_cyclonedx(self):
        component_bom_ref = "pkg:pypi/django@4.0.10"
        data = self.data / "cyclonedx/django-4.0.10-vulnerability.json"

        vulnerability_data = json.loads(data.read_text())
        results = output.vulnerability_as_cyclonedx(
            vulnerability_data, component_bom_ref
        )

        expected_location = self.data / "cyclonedx/django-4.0.10_as_cdx.json"
        results_as_json = results.as_json()

        # if True:
        #     expected_location.write_text(results_as_json)

        self.assertJSONEqual(results_as_json, expected_location.read_text())

    def test_scanpipe_pipes_outputs_to_cyclonedx(self, regen=FIXTURES_REGEN):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})

        project = Project.objects.get(name="asgiref")
        package = project.discoveredpackages.get(datasource_ids=["pypi_wheel_metadata"])

        package.other_license_expression_spdx = "Apache-2.0 AND LicenseRef-test"
        data = self.data / "cyclonedx/django-4.0.10-vulnerability.json"
        vulnerability_data = json.loads(data.read_text())
        package.affected_by_vulnerabilities = [vulnerability_data]
        package.save()

        with mock.patch("cyclonedx.model.bom.uuid4") as mock_uuid4:
            fake_uuid = uuid.UUID("b74fe5df-e965-415e-ba65-f38421a0695d")
            mock_uuid4.return_value = fake_uuid
            output_file = output.to_cyclonedx(project=project)

        self.assertIn(output_file.name, project.output_root)

        # Patch the tool version
        results_json = json.loads(output_file.read_text())
        results_json["metadata"]["tools"][0]["version"] = "0.0.0"
        results_json["metadata"]["timestamp"] = "2024-03-07T17:05:37.329061+00:00"
        results_json["vulnerabilities"][0]["bom-ref"] = "BomRef"
        results = json.dumps(results_json, indent=2)

        expected_location = self.data / "cyclonedx" / "asgiref-3.3.0.cdx.json"
        if regen:
            expected_location.write_text(results)

        self.assertJSONEqual(results, expected_location.read_text())

        output_file = output.to_cyclonedx(project=project, version="1.5")
        results_json = json.loads(output_file.read_text())
        self.assertEqual(
            "http://cyclonedx.org/schema/bom-1.5.schema.json", results_json["$schema"]
        )
        self.assertEqual("1.5", results_json["specVersion"])

    def test_scanpipe_pipes_outputs_get_cyclonedx_bom_dependency_tree(self):
        project = make_project(name="project")

        a = make_package(project, "pkg:type/a")
        b = make_package(project, "pkg:type/b")
        c = make_package(project, "pkg:type/c")
        make_package(project, "pkg:type/z")

        # Project -> A -> B -> C
        # Project -> Z
        make_dependency(project, for_package=a, resolved_to_package=b)
        make_dependency(project, for_package=b, resolved_to_package=c)

        with self.assertNumQueries(2):
            output_file = output.to_cyclonedx(project=project)
        results_json = json.loads(output_file.read_text())

        expected = [
            {
                "dependsOn": ["pkg:type/a", "pkg:type/b", "pkg:type/c", "pkg:type/z"],
                "ref": str(project.uuid),
            },
            {"dependsOn": ["pkg:type/b"], "ref": "pkg:type/a"},
            {"dependsOn": ["pkg:type/c"], "ref": "pkg:type/b"},
            {"ref": "pkg:type/c"},
            {"ref": "pkg:type/z"},
        ]
        self.assertEqual(expected, results_json["dependencies"])

    def test_scanpipe_pipes_outputs_get_cyclonedx_bom_package_uid_instances(self):
        project = make_project(name="project")
        make_package(project, "pkg:type/a", package_uid="pkg:type/a?uuid=1")
        make_package(project, "pkg:type/a", package_uid="pkg:type/a?uuid=2")

        output_file = output.to_cyclonedx(project=project)
        results_json = json.loads(output_file.read_text())
        self.assertEqual(2, len(results_json["components"]))

    def test_scanpipe_pipes_outputs_to_spdx(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")

        with self.assertNumQueries(9):
            output_file = output.to_spdx(project=project, include_files=True)
        self.assertIn(output_file.name, project.output_root)

        # Patch the `created` date and tool version
        results_json = json.loads(output_file.read_text())
        results_json["creationInfo"]["created"] = "2000-01-01T01:02:03Z"
        results_json["creationInfo"]["creators"] = ["Tool: ScanCode.io-31.0.0"]
        # Files ordering is system dependent, excluded for now
        results_json["files"] = []
        results = json.dumps(results_json, indent=2)

        expected_file = self.data / "asgiref" / "asgiref-3.3.0.spdx.json"
        self.assertResultsEqual(expected_file, results)

        # Make sure the output can be generated even if the work_directory was wiped
        shutil.rmtree(project.work_directory)
        output_file = output.to_spdx(project=project)
        self.assertIn(output_file.name, project.output_root)

    def test_scanpipe_pipes_outputs_to_spdx_extracted_licenses(self):
        project = make_project(name="Analysis")
        package_data = dict(package_data1)
        # ac3filter resolves as LicenseRef-scancode-ac3filter
        expression = "mit AND ac3filter"
        package_data["declared_license_expression"] = expression
        pipes.update_or_create_package(project, package_data)

        output_file = output.to_spdx(project=project, include_files=True)
        self.assertIn(output_file.name, project.output_root)

        results_json = json.loads(output_file.read_text())
        # mit is part of the SPDX license list, thus not in hasExtractedLicensingInfos
        self.assertEqual(1, len(results_json["hasExtractedLicensingInfos"]))
        license_infos = results_json["hasExtractedLicensingInfos"][0]
        self.assertEqual("LicenseRef-scancode-ac3filter", license_infos["licenseId"])
        self.assertEqual("AC3Filter License", license_infos["name"])
        expected = [
            "https://scancode-licensedb.aboutcode.org/ac3filter",
            "https://github.com/nexB/scancode-toolkit/tree/develop/src/"
            "licensedcode/data/licenses/ac3filter.LICENSE",
            "http://www.ac3filter.net/wiki/Download_AC3Filter",
            "http://ac3filter.net",
            "http://ac3filter.net/forum",
        ]
        self.assertEqual(expected, license_infos["seeAlsos"])
        self.assertTrue(license_infos["extractedText"].startswith("License:"))

    @mock.patch("uuid.uuid4")
    def test_scanpipe_pipes_outputs_to_spdx_dependencies(self, mock_uuid4):
        forced_uuid = "b74fe5df-e965-415e-ba65-f38421a0695d"
        mock_uuid4.return_value = forced_uuid
        project = Project.objects.create(name="Analysis", uuid=forced_uuid)

        a = make_package(
            project, "pkg:type/a", uuid="a83a60de-81bc-4bf4-b48c-dc78e0e658a9"
        )
        b = make_package(
            project, "pkg:type/b", uuid="81147701-285f-485c-ba36-9cd3742790b1"
        )
        # 1. Package resolved dependency
        make_dependency(
            project,
            for_package=a,
            resolved_to_package=b,
            uuid="3c3ac137-d573-43b0-be6b-3be2815d1c2f",
        )
        # 2. Package unresolved dependency
        make_dependency(
            project,
            for_package=b,
            dependency_uid="for_package_b",
            uuid="d0e1eab2-9b8b-449b-b9d1-12147ffdd8a8",
        )
        # 3. Project unresolved dependency
        unresolved_dependency = make_dependency(
            project,
            dependency_uid="unresolved",
            uuid="29fbe562-a191-44b4-88e8-a9678071ecee",
        )
        unresolved_dependency.set_package_url("pkg:type/unresolved")
        unresolved_dependency.save()
        # 4. Project package
        make_package(project, "pkg:type/z", uuid="e391c33e-d7d0-4a97-a3c3-e947375c53d5")

        self.assertEqual(3, project.discoveredpackages.count())
        self.assertEqual(3, project.discovereddependencies.count())

        output_file = output.to_spdx(project=project)
        results_json = json.loads(output_file.read_text())
        self.assertEqual(6, len(results_json["packages"]))
        self.assertEqual(6, len(results_json["relationships"]))

        # Patch the `created` date and tool version
        results_json["creationInfo"]["created"] = "2000-01-01T01:02:03Z"
        results_json["creationInfo"]["creators"] = ["Tool: ScanCode.io"]
        # Files ordering is system dependent, excluded for now
        results_json["files"] = []
        results = json.dumps(results_json, indent=2)

        expected_file = self.data / "spdx" / "dependencies.spdx.json"
        self.assertResultsEqual(expected_file, results)

    @mock.patch("uuid.uuid4")
    def test_scanpipe_pipes_outputs_to_spdx_get_inputs_as_spdx_packages(
        self, mock_uuid4
    ):
        forced_uuid = "b74fe5df-e965-415e-ba65-f38421a0695d"
        mock_uuid4.return_value = forced_uuid

        # 1. Input manually copied to Project's inputs
        project = make_project(name="Copied")
        copied_input = project.input_path / "input_filename"
        copied_input.touch()
        inputs_as_spdx_packages = output.get_inputs_as_spdx_packages(project)
        expected = [
            {
                "name": "input_filename",
                "SPDXID": f"SPDXRef-scancodeio-input-{forced_uuid}",
                "packageFileName": "input_filename",
                "licenseConcluded": "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "downloadLocation": "NOASSERTION",
                "filesAnalyzed": True,
                "licenseDeclared": "NOASSERTION",
            }
        ]
        inputs_spdx_as_dict = [package.as_dict() for package in inputs_as_spdx_packages]
        self.assertEqual(expected, inputs_spdx_as_dict)

        # 2. Input uploaded to Project's inputs
        project = make_project(name="Uploaded")
        uploaded_file = SimpleUploadedFile("filename.ext", content=b"content")
        input_source = project.add_upload(
            uploaded_file=uploaded_file,
        )
        inputs_as_spdx_packages = output.get_inputs_as_spdx_packages(project)
        expected = [
            {
                "name": "filename.ext",
                "SPDXID": f"SPDXRef-scancodeio-input-{input_source.uuid}",
                "packageFileName": "filename.ext",
                "licenseConcluded": "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "downloadLocation": "NOASSERTION",
                "filesAnalyzed": True,
                "licenseDeclared": "NOASSERTION",
            }
        ]
        inputs_spdx_as_dict = [package.as_dict() for package in inputs_as_spdx_packages]
        self.assertEqual(expected, inputs_spdx_as_dict)

        # 3. Fetched (download_url, purl, docker, git, ...)
        project = make_project(name="Fetched")
        input_from_download_url = project.add_input_source(
            download_url="https://download.url/archive.zip",
            filename="archive.zip",
        )
        input_from_purl = project.add_input_source(
            download_url="pkg:npm/dnd-core@7.0.2",
            filename="dnd-core-7.0.2.tgz",
        )
        input_from_docker = project.add_input_source(
            download_url="docker://registry.com/debian:10.9",
            filename="debian_10.9.tar",
        )
        inputs_as_spdx_packages = output.get_inputs_as_spdx_packages(project)
        inputs_spdx_as_dict = [package.as_dict() for package in inputs_as_spdx_packages]
        self.maxDiff = None
        expected = [
            {
                "name": "archive.zip",
                "SPDXID": f"SPDXRef-scancodeio-input-{input_from_download_url.uuid}",
                "downloadLocation": "https://download.url/archive.zip",
                "licenseConcluded": "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "filesAnalyzed": True,
                "packageFileName": "archive.zip",
                "licenseDeclared": "NOASSERTION",
            },
            {
                "name": "debian_10.9.tar",
                "SPDXID": f"SPDXRef-scancodeio-input-{input_from_docker.uuid}",
                "downloadLocation": "docker://registry.com/debian:10.9",
                "licenseConcluded": "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "filesAnalyzed": True,
                "packageFileName": "debian_10.9.tar",
                "licenseDeclared": "NOASSERTION",
            },
            {
                "name": "dnd-core-7.0.2.tgz",
                "SPDXID": f"SPDXRef-scancodeio-input-{input_from_purl.uuid}",
                "downloadLocation": "pkg:npm/dnd-core@7.0.2",
                "licenseConcluded": "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "filesAnalyzed": True,
                "packageFileName": "dnd-core-7.0.2.tgz",
                "licenseDeclared": "NOASSERTION",
            },
        ]
        self.assertEqual(expected, inputs_spdx_as_dict)

    def test_scanpipe_pipes_outputs_to_to_ort_package_list_yml(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")

        with self.assertNumQueries(2):
            output_file = output.to_ort_package_list_yml(project=project)
        self.assertIn(output_file.name, project.output_root)

        expected_file = self.data / "asgiref" / "asgiref-3.3.0.package-list.yml"
        self.assertResultsEqual(expected_file, output_file.read_text())

    def test_scanpipe_pipes_outputs_to_all_formats(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")

        with self.assertNumQueries(35):
            output_file = output.to_all_formats(project=project)

        self.assertEqual("asgiref_outputs.zip", output_file.name)

        with zipfile.ZipFile(output_file, "r") as zip_ref:
            zip_contents = zip_ref.namelist()
            file_count = len(zip_contents)

        expected_file_count = len(output.FORMAT_TO_FUNCTION_MAPPING)
        self.assertEqual(expected_file_count, file_count)

    def test_scanpipe_pipes_outputs_to_all_outputs(self):
        fixtures = self.data / "asgiref" / "asgiref-3.3.0_fixtures.json"
        call_command("loaddata", fixtures, **{"verbosity": 0})
        project = Project.objects.get(name="asgiref")

        with self.assertNumQueries(0):
            output_file = output.to_all_outputs(project=project)

        self.assertEqual("asgiref_outputs.zip", output_file.name)

        with zipfile.ZipFile(output_file, "r") as zip_ref:
            zip_contents = zip_ref.namelist()
            file_count = len(zip_contents)

        self.assertEqual(len(project.output_root), file_count)

    def test_scanpipe_pipes_outputs_make_unknown_license_object(self):
        licensing = get_licensing()
        parsed_expression = licensing.parse("some-unknown-license")

        self.assertEqual(1, len(parsed_expression.symbols))
        license_symbol = list(parsed_expression.symbols)[0]
        license_object = output.make_unknown_license_object(license_symbol)

        self.assertEqual("some-unknown-license", license_object.key)
        self.assertEqual(
            "LicenseRef-unknown-some-unknown-license", license_object.spdx_license_key
        )
        self.assertEqual(
            "ERROR: Unknown license key, no text available.", license_object.text
        )
        self.assertFalse(license_object.is_builtin)

    def test_scanpipe_pipes_outputs_get_package_expression_symbols(self):
        licensing = get_licensing()
        parsed_expression = licensing.parse("mit AND some-unknown-license")
        symbols = output.get_package_expression_symbols(parsed_expression)
        self.assertEqual(2, len(symbols))
        self.assertTrue(hasattr(symbols[0], "wrapped"))
        self.assertTrue(hasattr(symbols[1], "wrapped"))

    def test_scanpipe_pipes_outputs_get_expression_as_attribution_links(self):
        expression = "mit AND gpl-2.0 with classpath-exception-2.0"
        licensing = get_licensing()
        parsed_expression = licensing.parse(expression)
        rendered = output.get_expression_as_attribution_links(parsed_expression)
        expected = (
            '<a href="#license_gpl-2.0">GPL-2.0-only</a>'
            " WITH "
            '<a href="#license_classpath-exception-2.0">Classpath-exception-2.0</a>'
            " AND "
            '<a href="#license_mit">MIT</a>'
        )
        self.assertEqual(expected, rendered)

    def test_scanpipe_pipes_outputs_render_template(self):
        template_location = str(self.data / "outputs" / "render_me.html")
        template_string = Path(template_location).read_text()
        context = {"var": "value"}
        rendered = output.render_template(template_string, context)
        self.assertEqual("value", rendered)

    def test_scanpipe_pipes_outputs_render_template_file(self):
        template_location = str(self.data / "outputs" / "render_me.html")
        context = {"var": "value"}
        rendered = output.render_template_file(template_location, context)
        self.assertEqual("value", rendered)

    def test_scanpipe_pipes_outputs_get_attribution_template(self):
        project = make_project(name="Analysis")
        template_location = str(output.get_attribution_template(project))
        expected_location = "templates/scanpipe/attribution.html"
        self.assertTrue(template_location.endswith(expected_location))

        config_directory = make_config_directory(project)
        custom_template_dir = config_directory / "templates"
        custom_template_dir.mkdir(parents=True)
        custom_attribution_template = custom_template_dir / "attribution.html"
        custom_attribution_template.touch()

        template_location = str(output.get_attribution_template(project))
        expected_location = "codebase/.scancode/templates/attribution.html"
        self.assertTrue(template_location.endswith(expected_location))

    def test_scanpipe_pipes_outputs_get_package_data_for_attribution(self):
        project = make_project(name="Analysis")
        package_data = dict(package_data1)
        expression = "mit AND gpl-2.0 AND mit"
        package_data["declared_license_expression"] = expression
        package = pipes.update_or_create_package(project, package_data)

        data = output.get_package_data_for_attribution(package, get_licensing())
        self.assertEqual("pkg:deb/debian/adduser@3.118?arch=all", data["package_url"])
        expected = (
            '<a href="#license_gpl-2.0">GPL-2.0-only</a> '
            'AND <a href="#license_mit">MIT</a>'
        )
        self.assertEqual(expected, data["expression_links"])
        expected = ["mit", "gpl-2.0"]
        licenses = [license.key for license in data["licenses"]]
        self.assertEqual(sorted(expected), sorted(licenses))

    def test_scanpipe_pipes_outputs_to_attribution(self):
        project = make_project(name="Analysis")
        package_data = dict(package_data1)
        expression = "mit AND gpl-2.0 with classpath-exception-2.0 AND missing-unknown"
        package_data["declared_license_expression"] = expression
        package_data["notice_text"] = "Notice text"
        pipes.update_or_create_package(project, package_data)

        with self.assertNumQueries(2):
            output_file = output.to_attribution(project=project)

        expected_file = self.data / "outputs" / "expected_attribution.html"
        self.assertResultsEqual(expected_file, output_file.read_text())

        config_directory = make_config_directory(project)
        custom_template_dir = config_directory / "templates"
        custom_template_dir.mkdir(parents=True)
        custom_attribution_template = custom_template_dir / "attribution.html"
        custom_attribution_template.touch()
        custom_attribution_template.write_text("EMPTY_TEMPLATE")

        output_file = output.to_attribution(project=project)
        self.assertEqual("EMPTY_TEMPLATE", output_file.read_text())


class ScanPipeXLSXOutputPipesTest(TestCase):
    def test_add_xlsx_worksheet_does_truncates_long_strings_over_max_len(self):
        # This test verifies that we do not truncate long text silently

        test_dir = Path(tempfile.mkdtemp(prefix="scancode-io-test"))

        # The max length that Excel supports is 32,767 char per cell.
        # and 253 linefeed. This does not seem to be an absolute science
        # though.
        len_32760_string = "f" * 32760
        len_10_string = "0123456789"

        values = get_cell_texts(
            original_text=len_32760_string + len_10_string,
            test_dir=test_dir,
            workbook_name="long-text",
        )

        expected = [
            None,
            None,
            "foo",
            None,
            "xlsx_errors",
            None,
            "fffffffffffffffffffffffffffffffffffffffffff0123456",
            None,
            "32767 length to fit in an XLSX cell maximum length",
        ]

        for r, x in zip(values, expected):
            if r != x:
                self.assertEqual(r[-50:], x)

    def test_add_xlsx_worksheet_does_not_munge_long_strings_of_over_1024_lines(self):
        # This test verifies that we do not truncate long text silently

        test_dir = Path(tempfile.mkdtemp(prefix="scancode-io-test"))

        len_1025_lines = "\r\n".join(str(i) for i in range(1025)) + "abcd"

        values = get_cell_texts(
            original_text=len_1025_lines,
            test_dir=test_dir,
            workbook_name="long-text",
        )
        expected = [
            None,
            None,
            "foo",
            None,
            "xlsx_errors",
            None,
            "5\n1016\n1017\n1018\n1019\n1020\n1021\n1022\n1023\n1024abcd",
            None,
            "The value of: foo has been truncated from: 65476 to 32767 length "
            "to fit in an XLSL cell maximum length",
        ]

        for r, x in zip(values, expected):
            if r != x:
                self.assertEqual(r[-50:], x)

    def test__adapt_value_for_xlsx_does_adapt(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="foo",
            value="some simple value",
        )
        self.assertEqual(result, "some simple value")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_to_string(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="foo",
            value=12.4,
        )
        self.assertEqual(result, "12.4")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_crlf_to_lf(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="foo",
            value="some \r\nsimple \r\nvalue\r\n",
        )
        self.assertEqual(result, "some \nsimple \nvalue\n")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_description_and_keeps_only_5_lines(self):
        twenty_lines = "\r\n".join(str(i) for i in range(20))

        result, error = output._adapt_value_for_xlsx(
            fieldname="description", value=twenty_lines
        )
        self.assertEqual(result, "0\n1\n2\n3\n4")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_copyrights(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="copyrights",
            value=[{"copyright": "bar"}, {"copyright": "foo"}, {"copyright": "bar"}],
        )
        self.assertEqual(result, "bar\nfoo")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_authors(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="authors",
            value=[{"author": "bar"}, {"author": "foo"}, {"author": "bar"}],
        )
        self.assertEqual(result, "bar\nfoo")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_holders(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="holders",
            value=[{"holder": "bar"}, {"holder": "foo"}, {"holder": "bar"}],
        )
        self.assertEqual(result, "bar\nfoo")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_emails(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="emails", value=[{"email": "foo@bar.com"}]
        )
        self.assertEqual(result, "foo@bar.com")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_urls(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="urls", value=[{"url": "http://bar.com"}]
        )
        self.assertEqual(result, "http://bar.com")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_dicts(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="somename", value={"value1": "bar", "value2": "foo"}
        )
        self.assertEqual(result, "value1: bar\nvalue2: foo\n")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_lists(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="somename", value=["value1", "bar", "value2", "foo"]
        )
        self.assertEqual(result, "value1\nbar\nvalue2\nfoo")
        self.assertEqual(error, None)

    def test__adapt_value_for_xlsx_does_adapt_tuples_and_removes_dueps(self):
        result, error = output._adapt_value_for_xlsx(
            fieldname="somename", value=["value1", "bar", "value1", "foo"]
        )
        self.assertEqual(result, "value1\nbar\nfoo")
        self.assertEqual(error, None)

    def test_get_unique_licenses_returns_unique_and_preserve_order(self):
        @dataclass
        class Lic:
            key: str

        packages = [
            {
                "name": "foo",
                "purl": "pkg:generic/foo",
                "licenses": [Lic("foo"), Lic("bar"), Lic("foo")],
            }
        ]
        result = output.get_unique_licenses(packages)
        expected = [Lic("foo"), Lic("bar")]
        self.assertEqual(result, expected)

        packages = [
            {
                "name": "foo",
                "purl": "pkg:generic/foo",
                "licenses": [Lic("bar"), Lic("foo"), Lic("foo")],
            }
        ]
        result = output.get_unique_licenses(packages)
        expected = [Lic("bar"), Lic("foo")]
        self.assertEqual(result, expected)

        packages = [
            {
                "name": "foo",
                "purl": "pkg:generic/foo",
                "licenses": [Lic("bar")],
            }
        ]
        result = output.get_unique_licenses(packages)
        expected = [Lic("bar")]
        self.assertEqual(result, expected)

    def test_get_unique_licenses_does_not_fail_on_empties(self):
        packages = [
            {
                "name": "foo",
                "purl": "pkg:generic/foo",
                "licenses": [],
            }
        ]
        result = output.get_unique_licenses(packages)
        expected = []
        self.assertEqual(result, expected)

        packages = [
            {
                "name": "foo",
                "purl": "pkg:generic/foo",
                "licenses": None,
            }
        ]
        result = output.get_unique_licenses(packages)
        expected = []
        self.assertEqual(result, expected)

        packages = [{"name": "foo"}]
        result = output.get_unique_licenses(packages)
        expected = []
        self.assertEqual(result, expected)


def get_cell_texts(original_text, test_dir, workbook_name):
    """
    Create a workbook with a worksheet with a cell with ``original_text``
    and then extract, read and return the actual texts as a list.
    """

    class Row:
        """A mock Row with a single attribute storing a long string"""

        def __init__(self, foo):
            self.foo = foo

    rows = [Row(original_text)]

    output_file = test_dir / workbook_name
    with xlsxwriter.Workbook(str(output_file)) as workbook:
        output.add_xlsx_worksheet(
            workbook=workbook,
            worksheet_name="packages",
            rows=rows,
            fields=["foo"],
        )

    extract_dir = test_dir / "extracted"
    shutil.unpack_archive(
        filename=output_file,
        extract_dir=extract_dir,
        format="zip",
    )

    # This XML doc contains the strings stored in cells and has this shape:
    # <?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    # <sst     xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    #      count="2" uniqueCount="2">
    #   <si><t>foo</t></si>
    #   <si><t>f0123456789</t></si>
    # </sst>

    shared_strings = extract_dir / "xl" / "sharedStrings.xml"
    # Using lxml.etree.parse to parse untrusted XML data is known to be vulnerable
    # to XML attacks. This is not an issue here as we are parsing a properly crafted
    # test file, not a maliciously crafted one.
    sstet = etree.parse(str(shared_strings))  # noqa: S320
    # in our special case the text we care is the last element of the XML

    return [t.text for t in sstet.getroot().iter()]
